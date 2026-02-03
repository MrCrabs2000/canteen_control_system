from openpyxl import Workbook
from pathlib import Path
from datetime import datetime
from datebase.classes import Product, Dish, Menu, db, Info, AssociationUserMenus
import os


def export_payments():
    wb = Workbook()
    ws = wb.active
    ws.title = "Оплаты"

    ws['A1'] = 'Дата'
    ws['B1'] = 'Тип меню'
    ws['C1'] = 'Цена'
    ws['D1'] = 'Выдано'
    ws['E1'] = 'Оплата по абонементу'
    ws['F1'] = 'Разовый платёж'
    ws['G1'] = 'Прибыль'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['G'].width = 11
    row = 2

    menus = db.session.query(Menu).all()
    users = db.session.query(AssociationUserMenus).all()
    users_id = [i.user_id for i in users]
    info_user = {}
    if users_id:
        inform = db.session.query(Info).filter(Info.user_id.in_(users_id)).all()
        for info in inform:
            info_user[info.user_id] = info.abonement

    menu_abonement = {}
    menu_payment = {}
    for menu in menus:
        day_pay_amount, day_abonement_amount, day_amount = 0, 0, 0
        menu_user = []
        for user in users:
            if menu.id == user.menu_id:
                menu_user.append(user.user_id)
        abonement_amount = 0
        for user_id in menu_user:
            abonement = info_user.get(user_id)
            if abonement and abonement >= menu.date:
                abonement_amount += 1
        pay_amount = len(menu_user) - abonement_amount

        day_abonement_amount += abonement_amount
        day_pay_amount += pay_amount
        menu_abonement[menu.id] = day_abonement_amount
        menu_payment[menu.id] = day_pay_amount

    db.session.close()

    for menu in menus:
        if menu.get_amount > 0:
            ws[f'A{row}'] = menu.date
            ws[f'B{row}'] = menu.type
            ws[f'C{row}'] = menu.price
            ws[f'D{row}'] = menu.get_amount
            ws[f'E{row}'] = menu_abonement.get(menu.id)
            ws[f'F{row}'] = menu_payment.get(menu.id)
            ws[f'G{row}'] = menu.price * menu.get_amount
            row += 1

    os.makedirs('exel', exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = Path('exel') / f"payments_export_{timestamp}.xlsx"
    filepath = filepath.resolve()
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(filepath))
    return filepath


def export_attendance():
    wb = Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    menus = db.session.query(Menu).order_by(Menu.date.asc()).all()
    users = db.session.query(AssociationUserMenus).all()
    users_id = [i.user_id for i in users]
    classes_list = []
    user_classes = {}

    if users_id:
        inform = db.session.query(Info).filter(Info.user_id.in_(users_id)).all()
        for info in inform:
            stud_class = info.stud_class if info.stud_class else 'Не указан'
            if stud_class not in classes_list:
                classes_list.append(stud_class)
            user_classes[info.user_id] = stud_class

    db.session.close()

    ws['A1'] = 'Дата'
    ws['B1'] = 'Тип меню'
    ws['C1'] = 'Цена'
    ws['D1'] = 'Выдано'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11

    column = 5
    for stud_class in classes_list:
        ws.cell(row=1, column=column, value=stud_class)
        column += 1

    row = 2

    for menu in menus:
        if menu.get_amount > 0:
            ws[f'A{row}'] = menu.date
            ws[f'B{row}'] = menu.type
            ws[f'C{row}'] = menu.price
            ws[f'D{row}'] = menu.get_amount

            menu_classes = {}
            for stud_class in classes_list:
                menu_classes[stud_class] = 0

            for user in users:
                if menu.id == user.menu_id:
                    user_class = user_classes.get(user.user_id)
                    if user_class in menu_classes:
                        menu_classes[user_class] += 1

            column = 5
            for stud_class in classes_list:
                amount = menu_classes.get(stud_class, 0)
                ws.cell(row=row, column=column, value=amount)
                column += 1

        row += 1

    os.makedirs('exel', exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = Path('exel') / f"attendance_export_{timestamp}.xlsx"
    filepath = filepath.resolve()
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(filepath))
    return filepath


def export_products():
    products = db.session.query(Product).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Продукты"

    ws['A1'] = 'Название'
    ws['B1'] = 'Куплено'
    ws['C1'] = 'Потрачено'
    ws['D1'] = 'Единица'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 9

    row = 2

    for product in products:
        if product.buy_amount > 0:
            ws[f'A{row}'] = product.name
            ws[f'B{row}'] = product.buy_amount
            ws[f'C{row}'] = product.spend_amount
            ws[f'D{row}'] = product.measurement
            row += 1

    os.makedirs('exel', exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = Path('exel') / f"products_export_{timestamp}.xlsx"
    filepath = filepath.resolve()
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(filepath))
    return filepath


def export_dishes():
    dishes = db.session.query(Dish).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Блюда"

    ws['A1'] = 'Название'
    ws['B1'] = 'Категория'
    ws['C1'] = 'Готово на даннный момент'
    ws['D1'] = 'Приготовлено за всё время'
    ws['E1'] = 'Выдано за всё время'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 20
    row = 2

    for dish in dishes:
        if dish.cook_amount > 0:
            ws[f'A{row}'] = dish.name
            ws[f'B{row}'] = dish.category
            ws[f'C{row}'] = dish.amount
            ws[f'D{row}'] = dish.cook_amount
            ws[f'E{row}'] = dish.give_amount
            row += 1

    os.makedirs('exel', exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = Path('exel') / f"dishes_export_{timestamp}.xlsx"
    filepath = filepath.resolve()
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(filepath))
    return filepath


def export_menus():
    menus = db.session.query(Menu).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Меню"

    ws['A1'] = 'Дата'
    ws['B1'] = 'Тип'
    ws['C1'] = 'Выдано'
    ws['D1'] = 'Цена'
    ws['E1'] = 'Прибыль'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11

    row = 2

    for menu in menus:
        if menu.get_amount > 0:
            ws[f'A{row}'] = menu.date
            ws[f'B{row}'] = menu.type
            ws[f'C{row}'] = menu.get_amount
            ws[f'D{row}'] = menu.price
            ws[f'E{row}'] = menu.price * menu.get_amount
            row += 1

    os.makedirs('exel', exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = Path('exel') / f"menus_export_{timestamp}.xlsx"
    filepath = filepath.resolve()
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(filepath))
    return filepath